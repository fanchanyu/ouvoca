"""CSV / Excel 匯出服務 (v3.36)

電腦小白痛點：「我想把客戶清單給會計小姐」「匯出去寄 email」
解法：一個函式產一份 CSV（UTF-8 with BOM 給 Excel 不亂碼）或 .xlsx。

設計原則：
  • 純函式：接 db 回 bytes
  • CSV with BOM：Windows Excel 開啟不亂碼（utf-8-sig）
  • Excel：用 openpyxl，自動凍結首列 + 欄寬
  • 中文表頭 + 英文 fallback

可匯出的清單（覆蓋 80% 小廠日常）：
  customers / parts / suppliers / sales_orders / purchase_orders / inventory
"""
from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import Iterable

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession


# ════════════════════════════════════════════════════════════════════
# CSV helpers (BOM-prefixed so Excel detects UTF-8 correctly)
# ════════════════════════════════════════════════════════════════════

def _csv_bytes(rows: Iterable[list]) -> bytes:
    """Rows → UTF-8 BOM CSV bytes（Excel 開不亂碼）。"""
    buf = io.StringIO()
    w = csv.writer(buf, lineterminator="\n")
    for r in rows:
        w.writerow(r)
    # BOM 給 Windows Excel 識別 UTF-8
    return ("﻿" + buf.getvalue()).encode("utf-8")


def _xlsx_bytes(header: list[str], rows: list[list], sheet_name: str = "Sheet1") -> bytes:
    """產生 .xlsx — 首列凍結 + 自動欄寬。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel sheet 名上限 31 字

    # Header
    ws.append(header)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    for col_idx, _ in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Body
    for r in rows:
        ws.append(r)

    # 凍結首列 + 自動欄寬（粗估每欄字數）
    ws.freeze_panes = "A2"
    for col_idx, h in enumerate(header, 1):
        max_len = len(str(h))
        for r in rows:
            if col_idx - 1 < len(r):
                cell_val = r[col_idx - 1]
                if cell_val is not None:
                    # 中文字元寬 ≈ 2 ASCII
                    txt = str(cell_val)
                    width = sum(2 if ord(c) > 127 else 1 for c in txt)
                    if width > max_len:
                        max_len = width
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ════════════════════════════════════════════════════════════════════
# Customers
# ════════════════════════════════════════════════════════════════════

async def export_customers(db: AsyncSession, fmt: str = "csv") -> bytes:
    """匯出客戶清單。fmt = csv | xlsx。"""
    from app.models.crm_sales import Customer

    rows = (await db.execute(
        select(Customer).order_by(Customer.code)
    )).scalars().all()

    header = ["客戶編號", "客戶名稱", "等級", "聯絡人", "電話", "Email",
              "地址", "付款條件", "信用額度", "啟用"]
    body = [[
        c.code, c.name, c.grade or "", c.contact_person or "",
        c.contact_phone or "", c.contact_email or "",
        (c.address or "")[:200],
        c.payment_terms or "", c.credit_limit or 0,
        "是" if c.is_active else "否",
    ] for c in rows]

    return _xlsx_bytes(header, body, "Customers") if fmt == "xlsx" else _csv_bytes([header, *body])


# ════════════════════════════════════════════════════════════════════
# Parts
# ════════════════════════════════════════════════════════════════════

async def export_parts(db: AsyncSession, fmt: str = "csv") -> bytes:
    from app.models.inventory import Part

    rows = (await db.execute(
        select(Part).order_by(Part.part_no)
    )).scalars().all()

    header = ["料號", "品名", "類別", "單位", "規格",
              "最小庫存", "最大庫存", "安全庫存", "前置天數", "單位成本", "啟用"]
    body = [[
        p.part_no, p.name, p.category or "", p.unit or "",
        (p.specification or "")[:100],
        p.min_stock or 0, p.max_stock or 0, p.safety_stock or 0,
        p.lead_time_days or 0, p.unit_cost or 0,
        "是" if p.is_active else "否",
    ] for p in rows]

    return _xlsx_bytes(header, body, "Parts") if fmt == "xlsx" else _csv_bytes([header, *body])


# ════════════════════════════════════════════════════════════════════
# Suppliers
# ════════════════════════════════════════════════════════════════════

async def export_suppliers(db: AsyncSession, fmt: str = "csv") -> bytes:
    from app.models.purchase import Supplier

    rows = (await db.execute(
        select(Supplier).order_by(Supplier.code)
    )).scalars().all()

    header = ["供應商編號", "名稱", "層級", "聯絡人", "電話", "Email",
              "地址", "付款條件", "前置天數", "已核可", "啟用"]
    body = [[
        s.code, s.name, s.tier or "", s.contact_person or "",
        s.contact_phone or "", s.contact_email or "",
        (s.address or "")[:200],
        s.payment_terms or "", s.lead_time_days or 0,
        "是" if s.is_approved else "否",
        "是" if s.is_active else "否",
    ] for s in rows]

    return _xlsx_bytes(header, body, "Suppliers") if fmt == "xlsx" else _csv_bytes([header, *body])


# ════════════════════════════════════════════════════════════════════
# Sales Orders
# ════════════════════════════════════════════════════════════════════

async def export_sales_orders(db: AsyncSession, fmt: str = "csv") -> bytes:
    from app.models.crm_sales import SalesOrder, Customer

    so_rows = (await db.execute(
        select(SalesOrder).order_by(SalesOrder.order_date.desc())
    )).scalars().all()

    # Cache customer lookup
    cu_ids = {so.customer_id for so in so_rows if so.customer_id}
    cu_map = {}
    if cu_ids:
        cus = (await db.execute(
            select(Customer).where(Customer.id.in_(cu_ids))
        )).scalars().all()
        cu_map = {c.id: c for c in cus}

    header = ["訂單編號", "客戶編號", "客戶名稱", "狀態", "訂單日",
              "預計交期", "總金額", "幣別", "付款狀態"]
    body = []
    for so in so_rows:
        cu = cu_map.get(so.customer_id)
        body.append([
            so.so_no,
            cu.code if cu else "",
            cu.name if cu else "(未指定)",
            so.status or "",
            so.order_date.strftime("%Y-%m-%d") if so.order_date else "",
            so.requested_delivery_date.strftime("%Y-%m-%d") if so.requested_delivery_date else "",
            so.total_amount or 0,
            so.currency or "TWD",
            so.payment_status or "",
        ])

    return _xlsx_bytes(header, body, "SalesOrders") if fmt == "xlsx" else _csv_bytes([header, *body])


# ════════════════════════════════════════════════════════════════════
# Purchase Orders
# ════════════════════════════════════════════════════════════════════

async def export_purchase_orders(db: AsyncSession, fmt: str = "csv") -> bytes:
    from app.models.purchase import PurchaseOrder, Supplier

    po_rows = (await db.execute(
        select(PurchaseOrder).order_by(PurchaseOrder.order_date.desc())
    )).scalars().all()

    sup_ids = {po.supplier_id for po in po_rows if po.supplier_id}
    sup_map = {}
    if sup_ids:
        sups = (await db.execute(
            select(Supplier).where(Supplier.id.in_(sup_ids))
        )).scalars().all()
        sup_map = {s.id: s for s in sups}

    header = ["採購單號", "供應商編號", "供應商名稱", "狀態", "訂單日",
              "預計交期", "總金額", "幣別", "付款狀態"]
    body = []
    for po in po_rows:
        sup = sup_map.get(po.supplier_id)
        body.append([
            po.po_no,
            sup.code if sup else "",
            sup.name if sup else "(未指定)",
            po.status or "",
            po.order_date.strftime("%Y-%m-%d") if po.order_date else "",
            po.expected_delivery_date.strftime("%Y-%m-%d") if po.expected_delivery_date else "",
            po.total_amount or 0,
            po.currency or "TWD",
            po.payment_status or "",
        ])

    return _xlsx_bytes(header, body, "PurchaseOrders") if fmt == "xlsx" else _csv_bytes([header, *body])


# ════════════════════════════════════════════════════════════════════
# Inventory snapshot
# ════════════════════════════════════════════════════════════════════

async def export_inventory(db: AsyncSession, fmt: str = "csv") -> bytes:
    from app.models.inventory import Part, Inventory

    parts = (await db.execute(
        select(Part).order_by(Part.part_no)
    )).scalars().all()

    inv_map = {}
    inv_rows = (await db.execute(select(Inventory))).scalars().all()
    for inv in inv_rows:
        inv_map[inv.part_id] = inv

    header = ["料號", "品名", "單位", "現有量", "可用量",
              "保留量", "安全庫存", "最小庫存", "最大庫存", "需補貨"]
    body = []
    for p in parts:
        inv = inv_map.get(p.id)
        on_hand = inv.qty_on_hand if inv else 0
        available = inv.qty_available if inv else 0
        reserved = inv.qty_allocated if inv else 0
        need_reorder = "⚠️是" if on_hand < (p.safety_stock or 0) else ""
        body.append([
            p.part_no, p.name, p.unit or "",
            on_hand, available, reserved,
            p.safety_stock or 0, p.min_stock or 0, p.max_stock or 0,
            need_reorder,
        ])

    return _xlsx_bytes(header, body, "Inventory") if fmt == "xlsx" else _csv_bytes([header, *body])
