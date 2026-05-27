"""法規報表 generators（v3.10 Track C）。

3 個 generator：
  - render_form_401_html(form_data) → str (HTML, browser print → PDF)
  - render_ar_aging_xlsx(rows) → bytes (Excel)
  - render_inventory_monthly_xlsx(rows) → bytes (Excel)

設計原則：
  - 純 generator，不碰 DB（DB query 走 service / API 層）
  - 輸出格式精簡：HTML 用 inline CSS + @media print；Excel 用 openpyxl
  - 不依賴 reportlab / weasyprint（Windows 安裝痛）— 用 browser print PDF
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any


# ─────────────────────────────────────────────────────────
# 401 申報書 HTML（browser ctrl+P 印 PDF）
# ─────────────────────────────────────────────────────────

def render_form_401_html(form: dict[str, Any], company_name: str = "") -> str:
    """組 401 申報書 HTML（A4 print-friendly）。

    form: 來自 /api/tax/tw/401 的 dict（Form401Response.model_dump()）
    """
    period = form.get("period", "?")
    year = form.get("year", "?")
    period_no = form.get("period_no", "?")
    sales_taxable = float(form.get("sales_taxable", 0))
    sales_zero = float(form.get("sales_zero_rate", 0))
    sales_exempt = float(form.get("sales_exempt", 0))
    sales_total = float(form.get("sales_total", 0))
    output_tax = float(form.get("output_tax", 0))
    input_tax_g = float(form.get("input_tax_general", 0))
    input_tax_f = float(form.get("input_tax_fixed_asset", 0))
    tax_payable = float(form.get("tax_payable", 0))
    generated = form.get("generated_at", "")

    return f"""<!DOCTYPE html>
<html lang="zh-TW"><head>
<meta charset="utf-8">
<title>401 營業稅申報書 — {period}</title>
<style>
@media print {{
  @page {{ size: A4; margin: 1.5cm 1cm; }}
  .no-print {{ display: none; }}
}}
body {{
  font-family: 'Microsoft JhengHei', 'PingFang TC', sans-serif;
  max-width: 21cm; margin: 0 auto; padding: 1cm; color: #1f2937;
}}
.header {{ text-align: center; margin-bottom: 20px; }}
.header h1 {{ margin: 0; font-size: 22pt; }}
.header .sub {{ font-size: 11pt; color: #6b7280; margin-top: 4px; }}
.meta {{ display: flex; justify-content: space-between; font-size: 10pt; margin: 16px 0; }}
table {{ width: 100%; border-collapse: collapse; margin: 12px 0; }}
th, td {{ border: 1px solid #9ca3af; padding: 8px 10px; font-size: 11pt; }}
th {{ background: #f3f4f6; text-align: left; }}
td.amt {{ text-align: right; font-family: monospace; }}
.section-title {{
  background: #1e40af; color: white; padding: 6px 10px;
  font-weight: bold; margin-top: 18px; font-size: 12pt;
}}
.tax-payable {{ background: #fef3c7; font-weight: bold; }}
.footer {{ margin-top: 30px; font-size: 9pt; color: #6b7280; text-align: center; }}
.print-btn {{
  position: fixed; top: 16px; right: 16px;
  padding: 8px 16px; background: #1e40af; color: white;
  border: none; border-radius: 4px; cursor: pointer; font-size: 13pt;
}}
</style></head><body>

<button class="print-btn no-print" onclick="window.print()">🖨 列印 / 存 PDF</button>

<div class="header">
  <h1>營業人銷售額與稅額申報書（401）</h1>
  <div class="sub">VAT/Business Tax Form 401 — Bi-monthly Filing</div>
</div>

<div class="meta">
  <div><strong>所屬年度期別：</strong>{year} 年 第 {period_no} 期（{period}）</div>
  <div><strong>申報營業人：</strong>{company_name or '（未設定 — 請於 settings 填入）'}</div>
</div>

<div class="section-title">一、銷售額</div>
<table>
  <tr><th width="60%">項目</th><th>金額（新台幣 $）</th></tr>
  <tr><td>應稅銷售額（5%）</td><td class="amt">{sales_taxable:,.0f}</td></tr>
  <tr><td>零稅率銷售額</td><td class="amt">{sales_zero:,.0f}</td></tr>
  <tr><td>免稅銷售額</td><td class="amt">{sales_exempt:,.0f}</td></tr>
  <tr><td><strong>合計</strong></td><td class="amt"><strong>{sales_total:,.0f}</strong></td></tr>
</table>

<div class="section-title">二、銷項稅額</div>
<table>
  <tr><th width="60%">項目</th><th>稅額</th></tr>
  <tr><td>應稅銷售額 × 5%</td><td class="amt">{output_tax:,.0f}</td></tr>
</table>

<div class="section-title">三、進項稅額</div>
<table>
  <tr><th width="60%">項目</th><th>稅額</th></tr>
  <tr><td>進項稅額（一般）</td><td class="amt">{input_tax_g:,.0f}</td></tr>
  <tr><td>進項稅額（固定資產）</td><td class="amt">{input_tax_f:,.0f}</td></tr>
  <tr><td><strong>合計可扣抵</strong></td><td class="amt"><strong>{(input_tax_g + input_tax_f):,.0f}</strong></td></tr>
</table>

<div class="section-title">四、應納稅額（或可退）</div>
<table>
  <tr class="tax-payable">
    <td>本期應納稅額（銷項 - 進項）</td>
    <td class="amt">{tax_payable:,.0f}</td>
  </tr>
</table>

<div class="footer">
  本表由 LLM-ERP v3.10 自動產生 · 產生時間 {generated}<br>
  ⚠️ 本表為輔助計算工具；正式申報請以財政部電子申報軟體為準。
</div>
</body></html>"""


# ─────────────────────────────────────────────────────────
# 應收帳齡表 Excel
# ─────────────────────────────────────────────────────────

def render_ar_aging_xlsx(rows: list[dict], generated_at: str = "") -> bytes:
    """應收帳齡 Excel。

    rows: [{
      "invoice_no": str, "customer_name": str, "amount": float,
      "paid_amount": float, "due_date": str, "aging_days": int, "status": str,
    }, ...]
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "應收帳齡表"

    title_font = Font(name="Microsoft JhengHei", size=14, bold=True, color="FFFFFF")
    header_font = Font(name="Microsoft JhengHei", size=11, bold=True)
    money_font = Font(name="Consolas", size=10)
    title_fill = PatternFill("solid", fgColor="1E40AF")
    header_fill = PatternFill("solid", fgColor="E5E7EB")
    overdue_fill = PatternFill("solid", fgColor="FEE2E2")

    # Title row
    ws.merge_cells("A1:H1")
    ws["A1"] = f"應收帳齡分析表（生成時間 {generated_at}）"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Headers
    headers = ["發票號", "客戶", "應收金額", "已收金額", "未收餘額",
               "到期日", "帳齡（天）", "狀態"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    # Data rows
    total_amount = 0.0
    total_paid = 0.0
    total_outstanding = 0.0
    aging_buckets = {"0-30": 0.0, "31-60": 0.0, "61-90": 0.0, "90+": 0.0}

    for i, r in enumerate(rows, start=3):
        amt = float(r.get("amount", 0))
        paid = float(r.get("paid_amount", 0))
        outstanding = amt - paid
        days = int(r.get("aging_days", 0))

        ws.cell(row=i, column=1, value=r.get("invoice_no", ""))
        ws.cell(row=i, column=2, value=r.get("customer_name", "") or r.get("customer_id", ""))
        ws.cell(row=i, column=3, value=amt).font = money_font
        ws.cell(row=i, column=4, value=paid).font = money_font
        ws.cell(row=i, column=5, value=outstanding).font = money_font
        ws.cell(row=i, column=6, value=str(r.get("due_date", "")))
        ws.cell(row=i, column=7, value=days).alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=8, value=r.get("status", ""))

        for col in (3, 4, 5):
            ws.cell(row=i, column=col).number_format = "#,##0"

        # Highlight overdue rows
        if days > 0 and r.get("status") != "paid":
            for col in range(1, 9):
                ws.cell(row=i, column=col).fill = overdue_fill

        total_amount += amt
        total_paid += paid
        total_outstanding += outstanding
        if days <= 30:
            aging_buckets["0-30"] += outstanding
        elif days <= 60:
            aging_buckets["31-60"] += outstanding
        elif days <= 90:
            aging_buckets["61-90"] += outstanding
        else:
            aging_buckets["90+"] += outstanding

    # Totals row
    n = len(rows) + 3
    ws.cell(row=n, column=1, value="合計").font = header_font
    ws.cell(row=n, column=3, value=total_amount).font = header_font
    ws.cell(row=n, column=4, value=total_paid).font = header_font
    ws.cell(row=n, column=5, value=total_outstanding).font = header_font
    for col in (3, 4, 5):
        ws.cell(row=n, column=col).number_format = "#,##0"
        ws.cell(row=n, column=col).fill = header_fill

    # Aging summary 在右下角
    summary_start = n + 3
    ws.cell(row=summary_start, column=1, value="帳齡彙總").font = title_font
    ws.cell(row=summary_start, column=1).fill = title_fill
    for i, (bucket, amt) in enumerate(aging_buckets.items(), 1):
        ws.cell(row=summary_start + i, column=1, value=bucket)
        ws.cell(row=summary_start + i, column=2, value=amt).number_format = "#,##0"

    # Column widths
    widths = [16, 24, 14, 14, 14, 14, 12, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────
# 庫存月報 Excel
# ─────────────────────────────────────────────────────────

def render_inventory_monthly_xlsx(
    rows: list[dict],
    period_label: str = "",
    include_cost: bool = True,
) -> bytes:
    """庫存月報 Excel。

    rows: [{
      "part_no": str, "name": str, "category": str, "unit": str,
      "qty_on_hand": float, "qty_available": float, "safety_stock": float,
      "unit_cost": float, "value": float,  # 帳面金額
    }, ...]

    F-4：include_cost=False 時不輸出「單位成本」「帳面金額」兩欄（防財務外洩）。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "庫存月報表"

    title_font = Font(name="Microsoft JhengHei", size=14, bold=True, color="FFFFFF")
    header_font = Font(name="Microsoft JhengHei", size=11, bold=True)
    title_fill = PatternFill("solid", fgColor="1E40AF")
    header_fill = PatternFill("solid", fgColor="E5E7EB")
    low_fill = PatternFill("solid", fgColor="FEF3C7")  # 低於安全庫存

    n_cols = 9 if include_cost else 7
    merge_range = f"A1:{get_column_letter(n_cols)}1"
    ws.merge_cells(merge_range)
    ws["A1"] = f"庫存月報表 — {period_label}"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["料號", "名稱", "類別", "單位", "在手量", "可用量", "安全庫存"]
    if include_cost:
        headers += ["單位成本", "帳面金額"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    total_value = 0.0
    low_count = 0

    for i, r in enumerate(rows, start=3):
        qty_on_hand = float(r.get("qty_on_hand", 0))
        qty_available = float(r.get("qty_available", 0))
        safety_stock = float(r.get("safety_stock", 0))

        ws.cell(row=i, column=1, value=r.get("part_no", ""))
        ws.cell(row=i, column=2, value=r.get("name", ""))
        ws.cell(row=i, column=3, value=r.get("category", ""))
        ws.cell(row=i, column=4, value=r.get("unit", ""))
        ws.cell(row=i, column=5, value=qty_on_hand)
        ws.cell(row=i, column=6, value=qty_available)
        ws.cell(row=i, column=7, value=safety_stock)

        if include_cost:
            unit_cost = float(r.get("unit_cost", 0))
            value = float(r.get("value", qty_on_hand * unit_cost))
            ws.cell(row=i, column=8, value=unit_cost).number_format = "#,##0.00"
            ws.cell(row=i, column=9, value=value).number_format = "#,##0"
            total_value += value

        if safety_stock > 0 and qty_available < safety_stock:
            for col in range(1, n_cols + 1):
                ws.cell(row=i, column=col).fill = low_fill
            low_count += 1

    # Totals
    n = len(rows) + 3
    ws.cell(row=n, column=1, value="合計").font = header_font
    if include_cost:
        ws.cell(row=n, column=9, value=total_value).font = header_font
        ws.cell(row=n, column=9).number_format = "#,##0"
        ws.cell(row=n, column=9).fill = header_fill

    # Notes
    ws.cell(row=n + 2, column=1, value=f"⚠️ 低於安全庫存品項：{low_count} 筆（黃色標示）").font = header_font

    widths = [16, 30, 14, 8, 12, 12, 12]
    if include_cost:
        widths += [12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
